"""XLSX / XLS / CSV spreadsheet parser using openpyxl.

Each worksheet is treated as a separate page and tables are extracted
from the structured grid.
"""

from __future__ import annotations

import csv
import io
import logging
from pathlib import Path
from typing import Any

from app.parsers.base import (
    BaseParser,
    ExtractedPage,
    ExtractedTable,
    ParseResult,
)

logger = logging.getLogger(__name__)


class XlsxParser(BaseParser):
    """Parse Excel spreadsheets and CSV files."""

    @property
    def supported_mimes(self) -> set[str]:
        return {
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
            "text/csv",
            "application/csv",
        }

    @property
    def supported_extensions(self) -> set[str]:
        return {".xlsx", ".xls", ".csv"}

    async def parse(self, file_path: Path) -> ParseResult:
        ext = file_path.suffix.lower()
        if ext == ".csv":
            return await self._parse_csv(file_path)
        return await self._parse_xlsx(file_path)

    async def _parse_xlsx(self, file_path: Path) -> ParseResult:
        from openpyxl import load_workbook

        result = ParseResult()

        try:
            wb = load_workbook(
                str(file_path), read_only=True, data_only=True
            )
        except Exception as exc:
            logger.error("Failed to open XLSX %s: %s", file_path, exc)
            result.warnings.append(f"Failed to open: {exc}")
            return result

        result.metadata = {
            "sheet_names": wb.sheetnames,
            "sheet_count": len(wb.sheetnames),
        }

        table_idx = 0

        for sheet_num, sheet_name in enumerate(wb.sheetnames, start=1):
            ws = wb[sheet_name]

            # Read all rows as strings
            all_rows: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                str_row = [
                    str(cell).strip() if cell is not None else ""
                    for cell in row
                ]
                all_rows.append(str_row)

            if not all_rows:
                continue

            # Build text representation
            text_lines: list[str] = [f"[Sheet: {sheet_name}]"]
            for row in all_rows:
                line = " | ".join(cell for cell in row if cell)
                if line.strip():
                    text_lines.append(line)

            result.pages.append(
                ExtractedPage(
                    page_number=sheet_num,
                    text="\n".join(text_lines),
                    section=sheet_name,
                )
            )

            # Treat entire sheet as a table if it has header-like first row
            if len(all_rows) >= 2:
                headers = all_rows[0]
                rows = all_rows[1:]

                # Filter out completely empty rows
                rows = [r for r in rows if any(cell for cell in r)]

                if rows:
                    result.tables.append(
                        ExtractedTable(
                            table_index=table_idx,
                            headers=headers,
                            rows=rows,
                            page_number=sheet_num,
                            caption=sheet_name,
                        )
                    )
                    table_idx += 1

        wb.close()
        result.page_count = len(result.pages)
        return result

    async def _parse_csv(self, file_path: Path) -> ParseResult:
        import chardet

        result = ParseResult()

        # Detect encoding
        raw = file_path.read_bytes()
        detected = chardet.detect(raw)
        encoding = detected.get("encoding", "utf-8") or "utf-8"

        try:
            text = raw.decode(encoding)
        except (UnicodeDecodeError, LookupError):
            text = raw.decode("utf-8", errors="replace")
            result.warnings.append(
                f"Encoding detection failed, fell back to UTF-8"
            )

        reader = csv.reader(io.StringIO(text))
        all_rows = list(reader)

        if not all_rows:
            return result

        result.pages.append(
            ExtractedPage(
                page_number=1,
                text=text,
            )
        )
        result.page_count = 1

        headers = all_rows[0]
        rows = [r for r in all_rows[1:] if any(cell.strip() for cell in r)]
        if rows:
            result.tables.append(
                ExtractedTable(
                    table_index=0,
                    headers=headers,
                    rows=rows,
                    page_number=1,
                    caption=file_path.stem,
                )
            )

        return result
